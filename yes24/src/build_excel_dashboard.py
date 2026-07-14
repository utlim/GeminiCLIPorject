"""
프로젝트명: YES24 컴퓨터/IT 베스트셀러 데이터 엑셀 대시보드 및 자동화 요약 분석기
파일 역할: 원천 데이터(bestsellers.csv)를 로드하여 전처리한 후, 데이터 요약 수식과 차트가 포함된 고화질의 엑셀 대시보드 파일(Bestsellers_Dashboard.xlsx)을 빌드합니다.
주요 기능:
  - pandas를 통한 csv 데이터 파싱 및 파생변수(할인율 숫자형, 적립 포인트 숫자형, 출판년월) 생성
  - openpyxl을 사용한 2개 시트(Dashboard, Data) 구조 생성
  - Data 시트: 전처리 완료된 1,000건의 행 데이터 탑재 및 수식 보조용 열 추가
  - Dashboard 시트: 
    * 핵심 KPI 카드 6개 배치 (총 도서 수, 평균 정가, 평균 판매가, 평균 할인율, 평균 판매지수, 실질 평균 평점) - 모두 SUM/AVERAGE/AVERAGEIF 등 엑셀 수식 활용
    * Top 10 출판사 요약 표 (COUNTIF, AVERAGEIF, AVERAGEIFS 수식 적용)
    * 가격대별 도서 분포 요약 표 (COUNTIFS, AVERAGEIFS 수식 적용)
    * 할인율별 도서 성과 요약 표 (COUNTIF, AVERAGEIF 수식 적용)
    * 엑셀 임베디드 바 차트 2종 생성 (출판사별 도서 수, 가격대별 도서 수)
  - Arial 폰트 적용, 셀 너비 최적화, 세련된 네이비/블루 컬러 팔레트 적용
작성자: Antigravity AI
생성일: 2026-07-13
"""

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def createExcelDashboard(csvPath: str, outputPath: str):
    """
    CSV 데이터를 읽어서 가공한 뒤 대시보드 엑셀 파일을 빌드합니다.
    """
    try:
        # 1. 데이터 로드 및 판다스 전처리
        if not os.path.exists(csvPath):
            raise FileNotFoundError(f"파일이 없습니다: {csvPath}")
            
        df = pd.read_csv(csvPath)
        
        # 할인율 정제 ('10%' -> 10.0)
        df['discount_rate_num'] = df['discount_rate'].str.replace('%', '', regex=False).astype(float)
        
        # 적립 포인트 정제 ('1,500원' -> 1500)
        df['point_num'] = df['point'].str.replace('원', '', regex=False).str.replace(',', '', regex=False).astype(float)
        
        # 출판 연도 및 월 정제
        def parseYearMonth(dateStr):
            try:
                y = int(re.search(r'(\d{4})년', str(dateStr)).group(1))
                m = int(re.search(r'(\d{2})월', str(dateStr)).group(1))
                return y, m
            except:
                return 2026, 1
                
        df['publish_year'] = df['publish_date'].apply(lambda x: parseYearMonth(x)[0])
        df['publish_month'] = df['publish_date'].apply(lambda x: parseYearMonth(x)[1])
        df['subtitle'] = df['subtitle'].fillna('부제 없음')
        df['author'] = df['author'].fillna('미상')

        # 2. openpyxl Workbook 생성
        wb = Workbook()
        
        # 첫 번째 시트는 Dashboard
        wsDash = wb.active
        wsDash.title = "Dashboard"
        wsDash.views.sheetView[0].showGridLines = True
        
        # 두 번째 시트는 Data
        wsData = wb.create_sheet(title="Data")
        wsData.views.sheetView[0].showGridLines = True
        
        # 3. Data 시트에 데이터 탑재
        headers = [
            "goods_no", "rank", "title", "subtitle", "author", "publisher", 
            "publish_date", "discount_rate", "sale_price", "original_price", 
            "point", "sale_index", "review_count", "rating",
            "discount_rate_num", "point_num", "publish_year", "publish_month"
        ]
        
        wsData.append(headers)
        for row in df[headers].itertuples(index=False):
            wsData.append(list(row))
            
        # Data 시트 스타일링 (헤더: 회색 셰이딩, 폰트 Arial 10pt bold)
        headerFill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        headerFont = Font(name="Arial", size=10, bold=True, color="000000")
        thinBorder = Border(
            left=Side(style='thin', color='CBD5E0'),
            right=Side(style='thin', color='CBD5E0'),
            top=Side(style='thin', color='CBD5E0'),
            bottom=Side(style='thin', color='CBD5E0')
        )
        
        for col_num in range(1, len(headers) + 1):
            cell = wsData.cell(row=1, column=col_num)
            cell.fill = headerFill
            cell.font = headerFont
            cell.alignment = Alignment(horizontal="center")
            
        # Data 데이터 영역 테두리 및 폰트
        dataFont = Font(name="Arial", size=10)
        for r in range(2, len(df) + 2):
            for c in range(1, len(headers) + 1):
                cell = wsData.cell(row=r, column=c)
                cell.font = dataFont
                cell.border = thinBorder
                # 수치 데이터 형식 지정
                if c in [9, 10, 16]: # 가격, 포인트
                    cell.number_format = '#,##0'
                elif c in [12, 13]: # 판매지수, 리뷰수
                    cell.number_format = '#,##0'
                elif c in [14]: # 평점
                    cell.number_format = '0.0'
                elif c in [15]: # 할인율
                    cell.number_format = '0'

        # 4. Dashboard 시트 디자인 및 수식 적용
        # 4-1. 타이틀 행 (행 1, A1:L1 병합)
        wsDash.merge_cells("A1:L1")
        titleCell = wsDash["A1"]
        titleCell.value = "YES24 IT/컴퓨터 베스트셀러 다차원 EDA 대시보드"
        titleCell.font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
        titleCell.fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # 딥네이비
        titleCell.alignment = Alignment(horizontal="center", vertical="center")
        wsDash.row_dimensions[1].height = 40
        
        # 4-2. 핵심 KPI 카드 영역 (행 3 ~ 행 5, 2x2 사이즈 카드)
        # 카드 스타일 정의
        kpiFill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
        kpiBorder = Border(
            left=Side(style='medium', color='2B6CB0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        labelFont = Font(name="Arial", size=9, bold=True, color="718096")
        valueFont = Font(name="Arial", size=14, bold=True, color="1A202C")
        
        # KPI 카드 정의 리스트 [ (Label, Formula, CellRange, Format) ]
        kpis = [
            ("총 도서 수", "=COUNTA(Data!C2:C1001)", "A3:B4", "0"),
            ("평균 정가", "=AVERAGE(Data!J2:J1001)", "C3:D4", "#,##0\"원\""),
            ("평균 판매가", "=AVERAGE(Data!I2:I1001)", "E3:F4", "#,##0\"원\""),
            ("평균 할인율", "=AVERAGE(Data!O2:O1001)", "G3:H4", "0.0\"%\""),
            ("평균 판매지수", "=AVERAGE(Data!L2:L1001)", "I3:J4", "#,##0"),
            ("실질 평균 평점", "=AVERAGEIF(Data!N2:N1001, \">0\")", "K3:L4", "0.00\"점\"")
        ]
        
        for label, formula, cr, num_format in kpis:
            # 병합
            wsDash.merge_cells(cr)
            start_cell = wsDash[cr.split(":")[0]]
            # 텍스트와 줄바꿈 처리를 위해 \n 사용 (카드 내부 줄바꿈 허용)
            start_cell.value = f"{label}\n\n{formula}"
            start_cell.font = valueFont
            start_cell.fill = kpiFill
            start_cell.border = kpiBorder
            start_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            start_cell.number_format = num_format
            
        wsDash.row_dimensions[3].height = 25
        wsDash.row_dimensions[4].height = 25
        
        # 4-3. 테이블 헤더/데이터 폰트 & 색상 정의
        thFill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid") # 블루
        thFont = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        tdBorder = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        totalFill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
        totalFont = Font(name="Arial", size=10, bold=True, color="000000")
        
        # 4-4. 테이블 1: Top 10 출판사 요약 표 (행 7 ~ 행 18)
        wsDash["A7"] = "출판사명"
        wsDash["B7"] = "도서 수"
        wsDash["C7"] = "평균 판매지수"
        wsDash["D7"] = "평균 평점"
        
        for col_letter in ["A", "B", "C", "D"]:
            wsDash[f"{col_letter}7"].fill = thFill
            wsDash[f"{col_letter}7"].font = thFont
            wsDash[f"{col_letter}7"].alignment = Alignment(horizontal="center")
            
        top_publishers = list(df['publisher'].value_counts().head(10).index)
        for idx, pub in enumerate(top_publishers):
            r = 8 + idx
            wsDash[f"A{r}"] = pub
            wsDash[f"B{r}"] = f"=COUNTIF(Data!$F$2:$F$1001, A{r})"
            wsDash[f"C{r}"] = f"=AVERAGEIF(Data!$F$2:$F$1001, A{r}, Data!$L$2:$L$1001)"
            wsDash[f"D{r}"] = f"=AVERAGEIFS(Data!$N$2:$N$1001, Data!$F$2:$F$1001, A{r}, Data!$N$2:$N$1001, \">0\")"
            
            for col_letter in ["A", "B", "C", "D"]:
                cell = wsDash[f"{col_letter}{r}"]
                cell.border = tdBorder
                cell.font = Font(name="Arial", size=10)
                if col_letter == "A":
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="right")
                    
                if col_letter == "B": cell.number_format = '#,##0'
                if col_letter == "C": cell.number_format = '#,##0'
                if col_letter == "D": cell.number_format = '0.00'
                
        # 18행: 합계/평균
        wsDash["A18"] = "합계/평균"
        wsDash["B18"] = "=SUM(B8:B17)"
        wsDash["C18"] = "=AVERAGE(C8:C17)"
        wsDash["D18"] = "=AVERAGE(D8:D17)"
        
        for col_letter in ["A", "B", "C", "D"]:
            cell = wsDash[f"{col_letter}18"]
            cell.fill = totalFill
            cell.font = totalFont
            cell.border = tdBorder
            if col_letter == "A":
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
            if col_letter == "B": cell.number_format = '#,##0'
            if col_letter == "C": cell.number_format = '#,##0'
            if col_letter == "D": cell.number_format = '0.00'
            
        # 4-5. 테이블 2: 가격대별 분포 요약 표 (행 21 ~ 행 26)
        wsDash["A21"] = "가격대 구간"
        wsDash["B21"] = "도서 수"
        wsDash["C21"] = "평균 판매지수"
        wsDash["D21"] = "평균 평점"
        
        for col_letter in ["A", "B", "C", "D"]:
            wsDash[f"{col_letter}21"].fill = thFill
            wsDash[f"{col_letter}21"].font = thFont
            wsDash[f"{col_letter}21"].alignment = Alignment(horizontal="center")
            
        price_ranges = [
            ("10,000원 이하", 0, 10000),
            ("10,000원 초과 ~ 20,000원 이하", 10000, 20000),
            ("20,000원 초과 ~ 30,000원 이하", 20000, 30000),
            ("30,000원 초과", 30000, 999999)
        ]
        
        for idx, (label, low, high) in enumerate(price_ranges):
            r = 22 + idx
            wsDash[f"A{r}"] = label
            if high < 999999:
                wsDash[f"B{r}"] = f"=COUNTIFS(Data!$I$2:$I$1001, \">{low}\", Data!$I$2:$I$1001, \"<={high}\")"
                wsDash[f"C{r}"] = f"=AVERAGEIFS(Data!$L$2:$L$1001, Data!$I$2:$I$1001, \">{low}\", Data!$I$2:$I$1001, \"<={high}\")"
                wsDash[f"D{r}"] = f"=AVERAGEIFS(Data!$N$2:$N$1001, Data!$I$2:$I$1001, \">{low}\", Data!$I$2:$I$1001, \"<={high}\", Data!$N$2:$N$1001, \">0\")"
            else:
                wsDash[f"B{r}"] = f"=COUNTIFS(Data!$I$2:$I$1001, \">{low}\")"
                wsDash[f"C{r}"] = f"=AVERAGEIFS(Data!$L$2:$L$1001, Data!$I$2:$I$1001, \">{low}\")"
                wsDash[f"D{r}"] = f"=AVERAGEIFS(Data!$N$2:$N$1001, Data!$I$2:$I$1001, \">{low}\", Data!$N$2:$N$1001, \">0\")"
                
            for col_letter in ["A", "B", "C", "D"]:
                cell = wsDash[f"{col_letter}{r}"]
                cell.border = tdBorder
                cell.font = Font(name="Arial", size=10)
                if col_letter == "A":
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="right")
                if col_letter == "B": cell.number_format = '#,##0'
                if col_letter == "C": cell.number_format = '#,##0'
                if col_letter == "D": cell.number_format = '0.00'
                
        # 26행: 합계/평균
        wsDash["A26"] = "합계/평균"
        wsDash["B26"] = "=SUM(B22:B25)"
        wsDash["C26"] = "=AVERAGE(C22:C25)"
        wsDash["D26"] = "=AVERAGE(D22:D25)"
        
        for col_letter in ["A", "B", "C", "D"]:
            cell = wsDash[f"{col_letter}26"]
            cell.fill = totalFill
            cell.font = totalFont
            cell.border = tdBorder
            if col_letter == "A":
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
            if col_letter == "B": cell.number_format = '#,##0'
            if col_letter == "C": cell.number_format = '#,##0'
            if col_letter == "D": cell.number_format = '0.00'

        # 4-6. 테이블 3: 할인율별 도서 요약 표 (행 29 ~ 행 34)
        wsDash["A29"] = "할인율"
        wsDash["B29"] = "도서 수"
        wsDash["C29"] = "평균 판매지수"
        wsDash["D29"] = "평균 평점"
        
        for col_letter in ["A", "B", "C", "D"]:
            wsDash[f"{col_letter}29"].fill = thFill
            wsDash[f"{col_letter}29"].font = thFont
            wsDash[f"{col_letter}29"].alignment = Alignment(horizontal="center")
            
        discounts = [10, 0, 5, 3]
        for idx, disc in enumerate(discounts):
            r = 30 + idx
            wsDash[f"A{r}"] = f"{disc}%"
            wsDash[f"B{r}"] = f"=COUNTIF(Data!$O$2:$O$1001, {disc})"
            wsDash[f"C{r}"] = f"=AVERAGEIF(Data!$O$2:$O$1001, {disc}, Data!$L$2:$L$1001)"
            wsDash[f"D{r}"] = f"=AVERAGEIFS(Data!$N$2:$N$1001, Data!$O$2:$O$1001, {disc}, Data!$N$2:$N$1001, \">0\")"
            
            for col_letter in ["A", "B", "C", "D"]:
                cell = wsDash[f"{col_letter}{r}"]
                cell.border = tdBorder
                cell.font = Font(name="Arial", size=10)
                if col_letter == "A":
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="right")
                if col_letter == "B": cell.number_format = '#,##0'
                if col_letter == "C": cell.number_format = '#,##0'
                if col_letter == "D": cell.number_format = '0.00'
                
        # 34행: 합계/평균
        wsDash["A34"] = "합계/평균"
        wsDash["B34"] = "=SUM(B30:B33)"
        wsDash["C34"] = "=AVERAGE(C30:C33)"
        wsDash["D34"] = "=AVERAGE(D30:D33)"
        
        for col_letter in ["A", "B", "C", "D"]:
            cell = wsDash[f"{col_letter}34"]
            cell.fill = totalFill
            cell.font = totalFont
            cell.border = tdBorder
            if col_letter == "A":
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
            if col_letter == "B": cell.number_format = '#,##0'
            if col_letter == "C": cell.number_format = '#,##0'
            if col_letter == "D": cell.number_format = '0.00'

        # 5. 엑셀 차트 2종 탑재
        # 5-1. 출판사별 도서 수 차트 (가로 막대 차트)
        chart1 = BarChart()
        chart1.type = "bar"
        chart1.style = 10
        chart1.title = "Top 10 출판사별 등록 도서 수"
        chart1.x_axis.title = "도서 수 (권)"
        chart1.y_axis.title = "출판사"
        
        data_ref = Reference(wsDash, min_col=2, min_row=7, max_row=17) # B7:B17 (헤더 포함)
        cats_ref = Reference(wsDash, min_col=1, min_row=8, max_row=17) # A8:A17
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        chart1.legend = None # 단일 계열이므로 범례 생략
        chart1.width = 16
        chart1.height = 10
        wsDash.add_chart(chart1, "F7") # F7 위치에 추가
        
        # 5-2. 가격대별 도서 수 차트 (세로 막대 차트)
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 11
        chart2.title = "가격구간별 베스트셀러 도서 수 분포"
        chart2.x_axis.title = "가격 구간"
        chart2.y_axis.title = "도서 수 (권)"
        
        data_ref2 = Reference(wsDash, min_col=2, min_row=21, max_row=25) # B21:B25
        cats_ref2 = Reference(wsDash, min_col=1, min_row=22, max_row=25) # A22:A25
        chart2.add_data(data_ref2, titles_from_data=True)
        chart2.set_categories(cats_ref2)
        chart2.legend = None
        chart2.width = 16
        chart2.height = 7
        wsDash.add_chart(chart2, "F21")
        
        # 6. 셀 폭 자동 조절
        # Dashboard 시트 폭 조절
        for col in ["A", "B", "C", "D"]:
            max_len = 0
            for r in range(1, 35):
                val = wsDash[f"{col}{r}"].value
                if val:
                    # 줄바꿈이 있는 KPI 카드는 줄 단위로 나누어 최대 길이 계산
                    lines = str(val).split('\n')
                    for line in lines:
                        if not line.startswith('='): # 수식은 제외하고 일반 텍스트 길이 측정
                            max_len = max(max_len, len(line))
            wsDash.column_dimensions[col].width = max(max_len + 5, 12)
            
        wsDash.column_dimensions["F"].width = 2 # 차트 여백용 얇은 열
        
        # Data 시트 폭 조절
        for col in wsData.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            wsData.column_dimensions[col_letter].width = max(min(max_len + 3, 40), 10)
            
        # 파일 저장
        os.makedirs(os.path.dirname(outputPath), exist_ok=True)
        wb.save(outputPath)
        print(f"[성공] 엑셀 대시보드 저장 완료: {outputPath}")
        
    except Exception as e:
        print(f"[오류] 엑셀 생성 중 예외 발생: {e}")

if __name__ == "__main__":
    createExcelDashboard(
        csvPath="yes24/data/bestsellers.csv",
        outputPath="yes24/reports/Bestsellers_Dashboard.xlsx"
    )
